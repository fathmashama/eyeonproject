import cv2
import face_recognition
import pickle
import os
import numpy as np
from datetime import datetime, timedelta
import csv
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# Load encodings
with open("FaceDataSheet/encodings.pkl", "rb") as file:
    data = pickle.load(file)
    known_encodings = data["encodings"]
    known_names = data["names"]

print("✅ Encodings loaded.")

# Load employee details from employee_info.csv
employee_info = {}
with open("employee_info.csv", mode='r') as csvfile:
    reader = csv.DictReader(csvfile)
    for row in reader:
        employee_info[row["Name"]] = {
            "EmployeeID": row["EmployeeID"],
            "Email": row["Email"]
        }

# Function to write to Excel (per-day file)
def write_to_excel(emp_id, name, email, date, time):
    excel_path = f"AttendanceRecords/{date}.xlsx"
    file_exists = os.path.exists(excel_path)

    if file_exists:
        wb = load_workbook(excel_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws.append(["EmployeeID", "Name", "Email", "Date", "Time"])

    ws.append([emp_id, name, email, date, time])

    # Auto-adjust column width
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(excel_path)

# Initialize webcam
video = cv2.VideoCapture(0)
print("📷 Webcam started. Press 'q' to quit.")

last_seen = {}
delay = timedelta(minutes=5)
attendance_dir = "AttendanceRecords"
os.makedirs(attendance_dir, exist_ok=True)

if not os.path.exists("attendance.csv"):
    with open("attendance.csv", "w") as f:
        f.write("EmployeeID,Name,Email,DateTime\n")

while True:
    ret, frame = video.read()
    if not ret:
        break

    if frame is None or frame.size == 0:
        continue

    # Handle 4-channel (BGRA) frames from some webcams
    if len(frame.shape) == 3 and frame.shape[2] == 4:
        frame = cv2.cvtColor(frame, cv2.COLOR_BGRA2BGR)

    rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

    # Ensure the image is contiguous uint8 (required by dlib)
    if rgb_frame.dtype != np.uint8:
        rgb_frame = rgb_frame.astype(np.uint8)
    if not rgb_frame.flags['C_CONTIGUOUS']:
        rgb_frame = np.ascontiguousarray(rgb_frame)

    face_locations = face_recognition.face_locations(rgb_frame)
    face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)

    for face_encoding, face_location in zip(face_encodings, face_locations):
        name = "Unknown"
        face_distances = face_recognition.face_distance(known_encodings, face_encoding)

        if len(face_distances) > 0:
            best_match_index = np.argmin(face_distances)
            min_distance = face_distances[best_match_index]

            if min_distance < 0.45:
                name = known_names[best_match_index]
                now = datetime.now()
                current_time = now.strftime("%H:%M:%S")
                current_date = now.strftime("%Y-%m-%d")
                timestamp_str = f"{name} | {current_date} | {current_time}"

                if name not in last_seen or now - last_seen[name] > delay:
                    print(f"✅ {timestamp_str}")
                    last_seen[name] = now

                    # Write to .txt file
                    txt_path = os.path.join(attendance_dir, f"{current_date}.txt")
                    with open(txt_path, "a") as f:
                        f.write(timestamp_str + "\n")

                    # Write to Excel file (per day)
                    emp_id = employee_info.get(name, {}).get("EmployeeID", "N/A")
                    email = employee_info.get(name, {}).get("Email", "N/A")
                    write_to_excel(emp_id, name, email, current_date, current_time)

                    now_datetime = now.strftime("%Y-%m-%d %H:%M:%S")
                    with open("attendance.csv", "a") as f:
                        f.write(f"{emp_id},{name},{email},{now_datetime}\n")

        # Draw box and name
        top, right, bottom, left = face_location
        cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
        cv2.putText(frame, name, (left, top - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (255, 255, 255), 2)

    cv2.imshow("Attendance System", frame)
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

video.release()
cv2.destroyAllWindows()
print("👋 Webcam closed.")
